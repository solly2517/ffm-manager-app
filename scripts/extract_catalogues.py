import json
import re
import subprocess
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path("/home/ubuntu")
UPLOADS = ROOT / "upload"
OUTPUT = Path("/home/ubuntu/ffm-manager-app/data/catalogue-import.json")
REPORT = Path("/home/ubuntu/ffm-manager-app/data/catalogue-import-report.json")
PDFS = [
    "aapPrimaryKneeCatalog-04Jul2021_NSh_05Jul2021.pdf",
    "aapRevisionKneeCatalog-04Jul2021_NSh_05Jul2021.pdf",
    "aapPrimaryHipCatalog-04Jul2021_NSh_05Jul2021.pdf",
    "aapRevisionHipCatalog-04Jul2021_NSh_05Jul2021.pdf",
    "Main-catalogue-2024.pdf",
    "TraumaPlatesandScrewsCatalog-April-25.pdf",
    "ExternalFixationSystem-FEB-2025.pdf",
    "SpineCatalogue-draft.pdf",
    "TraumaMasterCatalogue.pdf",
    "reverse-shoulder.pdf",
]

def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip(" ,;|-")

def key(value):
    return re.sub(r"[^a-z0-9]+", "", value.lower())

def source_manufacturer(filename):
    lowered = filename.lower()
    if lowered.startswith("aap"):
        return "AAP"
    if "medgal" in lowered or "main-catalogue" in lowered:
        return "MEDGAL"
    if "pitkar" in lowered or "externalfixation" in lowered:
        return "S.H. Pitkar Orthotools"
    if "spine" in lowered:
        return "Spine catalogue"
    if "trauma" in lowered:
        return "Trauma catalogue"
    if "shoulder" in lowered:
        return "Reverse shoulder catalogue"
    return "Supplied catalogue"

def add_item(items, name, code, manufacturer, source):
    name = clean(name)
    code = clean(code)
    if len(name) < 3 or len(name) > 220:
        return
    if name.lower() in {"description", "discription", "catalog", "catalogue", "table of contents", "contents"}:
        return
    if not any(char.isalpha() for char in name):
        return
    items.append({
        "name": name,
        "productCode": code or None,
        "manufacturer": clean(manufacturer) or None,
        "source": source,
    })

def extract_xlsx(items):
    workbook = load_workbook(UPLOADS / "Altamamproductsforuploading.xlsx", data_only=True, read_only=True)
    for sheet in workbook.worksheets:
        rows = [[clean(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
        header_index = None
        for index, row in enumerate(rows[:20]):
            joined = " ".join(value.lower() for value in row)
            if any(token in joined for token in ["description", "discription", "وصف", "article", "reference code", "code"]):
                header_index = index
                break
        if header_index is None:
            continue
        headers = [value.lower() for value in rows[header_index]]
        code_index = next((i for i, value in enumerate(headers) if any(token in value for token in ["code", "article", "reference"])), None)
        description_index = next((i for i, value in enumerate(headers) if any(token in value for token in ["description", "discription", "وصف"])), None)
        size_index = next((i for i, value in enumerate(headers) if "size" in value), None)
        for row in rows[header_index + 1:]:
            if not any(row):
                continue
            description = row[description_index] if description_index is not None and description_index < len(row) else ""
            code = row[code_index] if code_index is not None and code_index < len(row) else ""
            size = row[size_index] if size_index is not None and size_index < len(row) else ""
            if not description:
                candidates = [value for value in row if value and value != code]
                description = candidates[0] if candidates else ""
            if size and size.lower() not in description.lower():
                description = f"{description} — {size}"
            add_item(items, description, code, f"Altamam Excel — {sheet.title}", f"Altamamproductsforuploading.xlsx / {sheet.title}")

def extract_pdf(items, filename):
    path = UPLOADS / filename
    try:
        text = subprocess.run(["pdftotext", "-layout", str(path), "-"], check=True, text=True, stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, timeout=90).stdout
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired):
        return
    manufacturer = source_manufacturer(filename)
    code_first = re.compile(r"^(?P<code>(?=[A-Za-z0-9.-]*\d)[A-Za-z0-9]+(?:[.-][A-Za-z0-9]+)+)\s{2,}(?P<name>.{3,220})$")
    description_first = re.compile(r"^(?P<name>.{3,220}?)\s{2,}(?P<code>(?=[A-Za-z0-9. /-]*\d)[A-Za-z][A-Za-z0-9. /-]{2,70})$")
    for raw_line in text.splitlines():
        raw_table_line = raw_line.strip()
        line = clean(raw_table_line)
        lowered = line.lower()
        if not line or len(line) > 320 or lowered.startswith(("table of content", "contents", "page ")):
            continue
        match = code_first.match(raw_table_line) or description_first.match(raw_table_line)
        if not match or "catalog" in lowered or "system" in lowered and not re.search(r"\d", match.group("code")):
            continue
        name = clean(match.group("name"))
        product_code = clean(match.group("code"))
        if len(product_code) < 3 or not re.search(r"\d", product_code):
            continue
        add_item(items, name, product_code, manufacturer, filename)

def main():
    items = []
    extract_xlsx(items)
    for filename in PDFS:
        extract_pdf(items, filename)
    seen = set()
    normalized = []
    for item in items:
        identity = key(item["productCode"] or item["name"])
        if not identity or identity in seen:
            continue
        seen.add(identity)
        normalized.append(item)
    normalized.sort(key=lambda item: (item["manufacturer"] or "", item["name"].lower(), item["productCode"] or ""))
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
    report = {
        "total": len(normalized),
        "bySource": Counter(item["source"] for item in normalized),
        "byManufacturer": Counter(item["manufacturer"] or "Unspecified" for item in normalized),
        "withoutProductCode": sum(1 for item in normalized if not item["productCode"]),
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    main()
